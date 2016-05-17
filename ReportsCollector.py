#техническая почта
#wailt-467@kobzov-1219.iam.gserviceaccount.com

import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
import openpyxl
import datetime
import re
from DateParser import *
import numba
import os

class reportsCollector(object):
	

	def __init__(self):
		print("Подключаюcь к серверу...")
		#new oauth2client
		scope = ['https://spreadsheets.google.com/feeds']
		credentials = ServiceAccountCredentials.from_json_keyfile_name('Kobzov-a80e3a2bd7d0.json', scope)
		self.gc = gspread.authorize(credentials)

		
		self.allFiles=self.gc.openall()
		self.allValues = dict()
		self.allPolygraphy = dict()
		self.allKontakt = dict()
		self.currentMoment = str(datetime.datetime.today()).replace(":"," ").split(".")[0]
		
		print("Подключение прошло успешно!")
		
	def downloadFromGdrive(self,file):
		return self.gc.open(file.title)
		
	def getFileName(self,file):
		name = file.title
		return name.replace('.',' ').split(' ')
		
	def getAll(self):
		print("Выгружаю всю базу")
		print("Получаю список доступных файлов:")
		for file in self.allFiles:
			print("   ",file.title)

		print("\nСобираю данные планов и отчётов:")
		wb = Workbook()

		for file in self.allFiles:
			dateList = self.downloadFromGdrive(file).worksheet("Экспрес план-отчет").get_all_values()
			name = self.getFileName(file)

			title = dateList[8]
			ws = wb.active	
			ws.append(name[2:4])
			ws.append(title)
			
			for i in dateList:
				ws.append(i)
			ws.append([' '])
			print("   ",file.title)
		print('Информация собрана. Сохраняю её на диск')
		self.saveToXLSX(wb)

	def getDataAndAppendToReport(self,data,frame,wb):
		dateList = data.worksheet("Экспрес план-отчет").get_all_values()
		name = data.title
		self.allValues.update({(name,):dateList})
		
		title = dateList[8]
		ws = wb.active	
		ws.append(title)
		
		newList=[]
		for i in dateList:
			if inFrame(i[1],frame):
				newList+=[i]
		dateList = newList
				
		for i in dateList:
			ws.append(i)
		ws.append([' '])
		
	def getPolygraphyAndAppendToReport(self,data,wb):
		polyList = data.worksheet("Полиграфия").get_all_values()
		name = data.title
		self.allPolygraphy.update({(name,):polyList})
		ws = wb.active	
		for i in polyList:
			ws.append(i)
		ws.append([' '])
	
	def getKontaktAndAppendToReport(self,data,wb):
		konList = data.worksheet("База контрагентов").get_all_values()
		name = data.title
		self.allKontakt.update({(name,):konList})
		ws = wb.active	
		for i in konList:
			ws.append(i)
		ws.append([' '])
	
	def saveToXLSX(self,wb):
		lastMsg = 'ResultMass{}.xlsx'.format(str(self.currentMoment).replace(':',' '))
		wb.save(lastMsg)
	
	def saveToXLSXbyName(self,wb,name):
		date = self.currentMoment
		try:
			os.mkdir(os.getcwd()+"\\report"+date)
		except: 
			pass
		wb.save(str(os.getcwd())+"\\report"+date+"\\"+name+".xlsx")
	
	def collectReports(self,bd,td):
		print("Нужен отчёт с даты (д.м.г):",bd)
		print("По дату (д.м.г):",td)
		frame=[bd,td]
		print("Получаю список доступных файлов:")
		for file in self.allFiles:
			print("   ",file.title)
		print("\nСобираю данные планов и отчётов:")
		wbRep = Workbook()
		wbPoly = Workbook()
		wbKon = Workbook()
		for file in self.allFiles:
			print(file.title)
			data = self.downloadFromGdrive(file)
			
			self.getDataAndAppendToReport(data,frame,wbRep)
			self.getPolygraphyAndAppendToReport(data,wbPoly)
			self.getKontaktAndAppendToReport(data,wbKon)

		print('Информация собрана. Сохраняю её на диск')
		self.saveToXLSXbyName(wbRep,str(bd)+str(td)+"Reports")
		self.saveToXLSXbyName(wbPoly,str(bd)+str(td)+"Polygraphy")
		self.saveToXLSXbyName(wbKon,str(bd)+str(td)+"Kontakts")
		print("Информация сохранена успешно")

	def getListReport(self):
		wb = Workbook()
		ws = wb.active
		print("Выгружаю список доступных отчётов")
		for f in self.allFiles:
			ws.append(f.title.split(" "))
		self.saveToXLSXbyName(wb,"Available repots")
		print("Список сохранён успешно")	

	def getDate(self,admin,s,ws):
		x=0
		if s=="plan":
			x=1
		else: 
			x=5
		dateList=[]
		for lis in self.allValues[admin]:
			dateList.append(tuple(dateParse(lis[x])[0:2]))
		return set(dateList)
		
	def getVerdict(self,data):
		if data:
			return 'есть'
		else:
			return 'нет'	
		
	def createReport(self,bd,td):
		wb = openpyxl.load_workbook("init\\forma.xlsx")
		ws = wb[wb.worksheets[0].title]
		ws.append(["","",]+[str(bd[0])+"."+str(bd[1])]+[""]+[str(td[0])+"."+str(td[1])])	
		ws.append(["Фио администратора","Город","План","Отчёт","План","Контакты","Полиграфия"])
		
		for i in self.allValues:
			#вердикт по отчётам
			planYesterday=tuple(bd[0:2]) in self.getDate(i,"plan",ws)
			planToday= tuple(td[0:2]) in self.getDate(i,"plan",ws)
			reportYesterday = tuple(bd[0:2]) in self.getDate(i,"report",ws)
			
			py = self.getVerdict(planYesterday)
			pt = self.getVerdict(planToday)
			ry = self.getVerdict(reportYesterday)
			
			#вердикт по полиграфии	
			dateList=self.allPolygraphy[i]
			if len(self.allPolygraphy[i])>3:
				for j in dateList:
					if j[0]!='':
						lastPolyDate="обновлено "+j[0]
			else:
				lastPolyDate = "нет"
			
			#вердикт по контактам
			konList = self.allKontakt[i]
			if len(self.allKontakt[i])>3:
				for j in konList:
					if j[1]!='':
						lastKonDate="обновлено "+j[1]
			else:
				lastKonDate = "нет"
			
			#сборка отчёта
			name = i[0].split(" ")
			data =	[str(name[2])+" "+str(name[3]),name[1],py,ry,pt,lastKonDate,lastPolyDate]
			ws.append(data)
		self.saveToXLSXbyName(wb,str(bd)+str(td)+"dailyReport")